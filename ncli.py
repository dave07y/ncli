import argparse
import logging
import os
from openpyxl import load_workbook
from tqdm import tqdm
from copy import copy
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection

def setup_logging():
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    # Console Handler
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    console_format = logging.Formatter('%(message)s')
    ch.setFormatter(console_format)
    logger.addHandler(ch)

    # File Handler
    fh = logging.FileHandler('process.log', mode='w')
    fh.setLevel(logging.INFO)
    file_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    fh.setFormatter(file_format)
    logger.addHandler(fh)

    return logger

def remove_columns_from_workbook(input_file, output_file, columns_to_drop):
    wb = load_workbook(input_file)
    drop_values = set(columns_to_drop)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        if not header_row:
            continue
        header_cells = header_row[0]

        columns_to_remove = []
        for idx, cell in enumerate(header_cells, start=1):
            header_val = cell.value
            if str(header_val) in drop_values:
                columns_to_remove.append(idx)

        if not columns_to_remove:
            continue

        columns_to_remove.sort(reverse=True)

        for col_idx in columns_to_remove:
            max_col = ws.max_column
            for row_idx in range(1, ws.max_row + 1):
                for c in range(col_idx, max_col):
                    source_cell = ws.cell(row=row_idx, column=c+1)
                    target_cell = ws.cell(row=row_idx, column=c)

                    target_cell.value = source_cell.value
                    target_cell.number_format = source_cell.number_format
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.protection = copy(source_cell.protection)

                # Clear the last column after shifting
                last_cell = ws.cell(row=row_idx, column=max_col)
                last_cell.value = None
                last_cell.number_format = 'General'
                last_cell.font = Font()             # default font
                last_cell.border = Border()         # default border
                last_cell.fill = PatternFill()      # default fill
                last_cell.alignment = Alignment()   # default alignment
                last_cell.protection = Protection() # default protection

            ws.delete_cols(max_col)

    wb.save(output_file)

def drop_columns(args, logger):
    input_dir = args.input_dir
    output_dir = args.output_dir
    columns_file = args.columns_file

    logger.info("Starting the Excel Column Dropper process (openpyxl version).")
    logger.info(f"Input Directory: {input_dir}")
    logger.info(f"Output Directory: {output_dir}")

    os.makedirs(output_dir, exist_ok=True)

    if columns_file and os.path.exists(columns_file):
        with open(columns_file, 'r') as f:
            columns_to_drop = [line.strip() for line in f if line.strip()]
    else:
        columns_to_drop = []
    logger.info(f"Columns to drop: {columns_to_drop}")

    all_files = [f for f in os.listdir(input_dir) if f.lower().endswith(".xlsx")]
    if not all_files:
        logger.info("No Excel files found in the input directory.")
        return

    logger.info("Beginning file processing...")
    for filename in tqdm(all_files, desc="Processing files", unit="file"):
        input_path = os.path.join(input_dir, filename)
        output_path = os.path.join(output_dir, filename)

        try:
            remove_columns_from_workbook(input_path, output_path, columns_to_drop)
            logger.info(f"Processed: {filename} - Dropped columns: {columns_to_drop}")
        except Exception as e:
            logger.error(f"Failed to process {filename}: {e}")
            continue

    logger.info("Processing complete. Check the output directory and 'process.log' for details.")

def main():
    parser = argparse.ArgumentParser(prog="ncli", description="NCLI: A CLI tool for Excel processing with openpyxl.")
    subparsers = parser.add_subparsers(title="subcommands", dest="subcommand")

    drop_parser = subparsers.add_parser("dropcolumns", help="Drop specified columns from Excel files in a directory.")
    drop_parser.add_argument("--input-dir", required=True, help="Directory containing input Excel files")
    drop_parser.add_argument("--output-dir", default="processed", help="Directory to save processed files")
    drop_parser.add_argument("--columns-file", default="columns_to_drop.txt", help="File listing columns to drop")

    args = parser.parse_args()

    logger = setup_logging()

    if args.subcommand == "dropcolumns":
        drop_columns(args, logger)
    else:
        parser.print_help()

if __name__ == "__main__":
    main()
